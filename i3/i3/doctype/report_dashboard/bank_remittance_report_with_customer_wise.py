from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'Bank Remittance Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40 
    ws.column_dimensions['D'].width = 20
    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=5):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=4):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data = get_data(args)
    for d in data:
        ws.append(d)
    for rows in ws.iter_rows(min_row=3, max_row=len(d)+1, min_col=1, max_col=5):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def title1(args):
    data = ['']
    return data

@frappe.whitelist()
def header1(args):
    data = ['S.NO','Employee Code','Employee Name','Net Pay Amount']
    return data

@frappe.whitelist()
def get_data(args):
    if args.customer:

        data = []
        sa = frappe.get_all('Payslip', {'from_date': args.from_date, 'to_date': args.to_date,'customer':args.customer}, ['*'])
        for index, i in enumerate(sa, start=1):
           
            row=[index,i.employee,i.employee_name,i.total_net_pay]
            data.append(row)
        return data
        
        
    else:
        data = []
        sa = frappe.get_all('Salary Slips', {'from_date': args.from_date, 'to_date': args.to_date}, ['*'])
        for index, i in enumerate(sa, start=1):
            row=[index,i.employee,i.employee_name,i.total_net_pay]
            data.append(row)
        return data
        