from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'BOT Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    
    title_1=title1(args)
    ws.append(title_1)
    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    title_2=title2(args)
    ws.append(title_2)
    bold_font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = bold_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=6)


    title_3=title3(args)
    ws.append(title_3)
    bold_font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = bold_font
        ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=6)

    header_1=header1(args)
    ws.append(header_1)
    bold_font = Font(bold=True)
    for cell in ws["4:4"]:
        cell.font = bold_font
    for rows in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=6):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    bold_font = Font(bold=True)
    data = get_data(args)
    for d in data:
        ws.append(d)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def title1(args):
    data = ['Overall Consolidated Payroll Report']
    return data

def title2(args):
    month = datetime.strptime(str(args.from_date),'%Y-%m-%d')
    
    mon = str(str('Month:'+ month.strftime('%A')))
    data=[mon]
    return data
def title3(args):
    month = datetime.strptime(str(args.from_date),'%Y-%m-%d')
    mon = str(str('Year:'+ month.strftime('%Y')))
    data=[mon]
    return data
@frappe.whitelist()
def header1(args):
    data = ['S.NO','Employee Name','Employee Code','Client Name','BOT Days','BOT Net Amount',]
    return data

@frappe.whitelist()
def get_data(args):
    data = []
    sa = frappe.get_all('Payslip', {'from_date': args.from_date, 'to_date': args.to_date,}, ['*'])
    for index, i in enumerate(sa, start=1):
        v=i.bot_days
        w=i.working_days
        r=i.total_earnings
        bot_amt = (r / w) * v if w != 0 else 0

        row=[index,i.employee_name,i.employee,i.customer,i.bot_days,bot_amt]
        data.append(row)
    return data
